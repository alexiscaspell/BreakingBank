import csv
import io
import re
import uuid
from datetime import date, datetime, timedelta
from typing import Literal

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.constants.category_presets import preset_for_category_name
from app.models.account import Account
from app.models.category import Category
from app.models.label import Label, TransactionLabel
from app.models.transaction import Transaction
from app.utils.names import normalize_entity_name

HEADERS = [
    "Fecha y hora",
    "Categoría",
    "Cuenta",
    "Cantidad en la divisa de la cuenta",
    "Divisa de la cuenta",
    "Cantidad de la transacción en la divisa de la transacción",
    "Divisa de la transacción",
    "Etiquetas",
    "Comentario",
]

MONTHS_ES = [
    "enero",
    "febrero",
    "marzo",
    "abril",
    "mayo",
    "junio",
    "julio",
    "agosto",
    "septiembre",
    "octubre",
    "noviembre",
    "diciembre",
]

EXCEL_EPOCH = datetime(1899, 12, 30)


def month_bounds(year: int, month: int) -> tuple[date, date]:
    start = date(year, month, 1)
    if month == 12:
        end = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end = date(year, month + 1, 1) - timedelta(days=1)
    return start, end


def period_title(tx_type: str, year: int, month: int) -> str:
    kind = "gastos" if tx_type == "expense" else "ingresos"
    start_label = f"1 de {MONTHS_ES[month - 1]} de {year}"
    if month == 12:
        end_label = f"1 de enero de {year + 1}"
    else:
        end_label = f"1 de {MONTHS_ES[month]} de {year}"
    return f"Lista de {kind} para el período entre {start_label} y {end_label}"


def date_to_excel_serial(d: date) -> float:
    dt = datetime.combine(d, datetime.min.time())
    return float((dt - EXCEL_EPOCH).days)


def parse_excel_serial(value: float) -> date:
    whole = int(value)
    return (EXCEL_EPOCH + timedelta(days=whole)).date()


def parse_date_cell(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return parse_excel_serial(float(value))
    text = str(value).strip()
    if not text:
        return None
    if re.fullmatch(r"\d+(\.\d+)?", text):
        return parse_excel_serial(float(text))
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def parse_amount(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def split_labels(raw) -> list[str]:
    if raw is None or raw == "":
        return []
    return [part.strip() for part in str(raw).split(",") if part.strip()]


def row_to_values(
    txn_date: date,
    category_name: str,
    account_name: str,
    amount: float,
    currency: str,
    labels: list[str],
    comment: str | None,
) -> list:
    return [
        date_to_excel_serial(txn_date),
        category_name,
        account_name,
        amount,
        currency,
        "",
        "",
        ", ".join(labels),
        comment or "",
    ]


async def fetch_month_transactions(
    db: AsyncSession,
    group_id: str,
    tx_type: str,
    year: int,
    month: int,
) -> list[Transaction]:
    start, end = month_bounds(year, month)
    result = await db.execute(
        select(Transaction)
        .options(
            selectinload(Transaction.transaction_labels).selectinload(TransactionLabel.label),
            selectinload(Transaction.category),
            selectinload(Transaction.account),
        )
        .where(
            Transaction.group_id == group_id,
            Transaction.deleted_at.is_(None),
            Transaction.type == tx_type,
            Transaction.date >= start,
            Transaction.date <= end,
        )
        .order_by(Transaction.date.desc(), Transaction.created_at.desc())
    )
    return list(result.scalars().all())


def build_rows(
    transactions: list[Transaction],
    tx_type: str,
    year: int,
    month: int,
    currency: str,
) -> list[list]:
    rows: list[list] = [[period_title(tx_type, year, month)], HEADERS]
    for txn in transactions:
        labels = [tl.label.name for tl in txn.transaction_labels if tl.label]
        rows.append(
            row_to_values(
                txn.date,
                txn.category.name if txn.category else "",
                txn.account.name if txn.account else "",
                float(txn.amount),
                currency,
                labels,
                txn.comment,
            )
        )
    return rows


def rows_to_csv(rows: list[list]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    for row in rows:
        writer.writerow(row)
    return buf.getvalue().encode("utf-8-sig")


def rows_to_xlsx(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_sheet_rows(rows: list[tuple]) -> list[dict]:
    if len(rows) < 2:
        return []
    header_idx = None
    for i, row in enumerate(rows[:5]):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if cells[: len(HEADERS)] == HEADERS:
            header_idx = i
            break
    if header_idx is None:
        return []
    parsed: list[dict] = []
    for row in rows[header_idx + 1 :]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        cells = list(row) + [None] * (9 - len(row))
        amount = parse_amount(cells[3]) or parse_amount(cells[5])
        if amount is None:
            continue
        parsed.append(
            {
                "date": parse_date_cell(cells[0]),
                "category": str(cells[1]).strip() if cells[1] else "",
                "account": str(cells[2]).strip() if cells[2] else "",
                "amount": amount,
                "labels": split_labels(cells[7]),
                "comment": str(cells[8]).strip() if cells[8] else None,
            }
        )
    return parsed


def parse_csv_bytes(raw: bytes) -> list[dict]:
    text = raw.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    return parse_sheet_rows([tuple(row) for row in reader])


def parse_xlsx_bytes(raw: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active
    rows = [tuple(cell.value for cell in row) for row in ws.iter_rows()]
    return parse_sheet_rows(rows)


async def resolve_account(db: AsyncSession, group_id: str, user_id: str, name: str) -> Account:
    result = await db.execute(
        select(Account).where(
            Account.group_id == group_id,
            Account.deleted_at.is_(None),
            Account.name == name,
        )
    )
    acc = result.scalar_one_or_none()
    if acc:
        return acc
    acc = Account(
        id=str(uuid.uuid4()),
        client_id=str(uuid.uuid4()),
        user_id=user_id,
        group_id=group_id,
        name=name,
    )
    db.add(acc)
    await db.flush()
    return acc


async def find_category_by_name(
    db: AsyncSession, group_id: str, name: str, tx_type: str
) -> Category | None:
    norm = normalize_entity_name(name)
    if not norm:
        return None
    result = await db.execute(
        select(Category).where(
            Category.group_id == group_id,
            Category.deleted_at.is_(None),
            Category.type == tx_type,
        )
    )
    for cat in result.scalars().all():
        if normalize_entity_name(cat.name) == norm:
            return cat
    return None


async def resolve_category(
    db: AsyncSession, group_id: str, user_id: str, name: str, tx_type: str
) -> Category:
    cleaned = " ".join(str(name).strip().split())
    cat = await find_category_by_name(db, group_id, cleaned, tx_type)
    if cat:
        return cat
    preset = preset_for_category_name(cleaned, tx_type)
    if preset:
        display_name, color, icon_key = preset
    else:
        display_name, color, icon_key = cleaned, "#9E9E9E", "other"
    cat = Category(
        id=str(uuid.uuid4()),
        client_id=str(uuid.uuid4()),
        user_id=user_id,
        group_id=group_id,
        name=display_name,
        type=tx_type,
        color=color,
        icon_type="preset",
        icon_key=icon_key,
    )
    db.add(cat)
    await db.flush()
    return cat


async def resolve_label(db: AsyncSession, group_id: str, user_id: str, name: str) -> Label:
    result = await db.execute(
        select(Label).where(
            Label.group_id == group_id,
            Label.deleted_at.is_(None),
            Label.name == name,
        )
    )
    label = result.scalar_one_or_none()
    if label:
        return label
    label = Label(
        id=str(uuid.uuid4()),
        client_id=str(uuid.uuid4()),
        user_id=user_id,
        group_id=group_id,
        name=name,
    )
    db.add(label)
    await db.flush()
    return label


async def import_transactions(
    db: AsyncSession,
    group_id: str,
    user_id: str,
    tx_type: str,
    rows: list[dict],
) -> dict:
    created = 0
    skipped = 0
    errors: list[str] = []
    dates: list[date] = []

    for i, row in enumerate(rows, start=1):
        if not row.get("category") or not row.get("account"):
            skipped += 1
            errors.append(f"Fila {i}: falta categoría o cuenta")
            continue
        if not row.get("date"):
            skipped += 1
            errors.append(f"Fila {i}: fecha inválida")
            continue
        dates.append(row["date"])
        account = await resolve_account(db, group_id, user_id, row["account"])
        category = await resolve_category(db, group_id, user_id, row["category"], tx_type)
        txn = Transaction(
            id=str(uuid.uuid4()),
            client_id=str(uuid.uuid4()),
            user_id=user_id,
            group_id=group_id,
            account_id=account.id,
            category_id=category.id,
            type=tx_type,
            amount=row["amount"],
            date=row["date"],
            comment=row.get("comment") or None,
        )
        db.add(txn)
        await db.flush()
        for label_name in row.get("labels") or []:
            label = await resolve_label(db, group_id, user_id, label_name)
            db.add(TransactionLabel(transaction_id=txn.id, label_id=label.id))
        created += 1

    await db.commit()
    result = {"created": created, "skipped": skipped, "errors": errors}
    if dates:
        result["date_from"] = min(dates).isoformat()
        result["date_to"] = max(dates).isoformat()
    return result


async def export_month(
    db: AsyncSession,
    group_id: str,
    tx_type: Literal["expense", "income"],
    year: int,
    month: int,
    fmt: Literal["xlsx", "csv"],
    currency: str = "ARS",
) -> tuple[bytes, str]:
    txns = await fetch_month_transactions(db, group_id, tx_type, year, month)
    rows = build_rows(txns, tx_type, year, month, currency)
    stamp = datetime.now().strftime("%Y_%m_%d_%H_%M_%S_%f")
    if fmt == "csv":
        return rows_to_csv(rows), f"{stamp}.csv"
    return rows_to_xlsx(rows), f"{stamp}.xlsx"
